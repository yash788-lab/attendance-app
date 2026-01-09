from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from database import db
from models import Student, Attendance
from datetime import datetime, date, timedelta
from sqlalchemy import func
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

main = Blueprint('main', __name__)

@main.route('/')
def index():
    """Home page showing dashboard"""
    total_students = Student.query.count()
    today = date.today()
    today_attendance = Attendance.query.filter_by(date=today).count()
    
    # Get recent attendance records
    recent_records = Attendance.query.order_by(Attendance.created_at.desc()).limit(10).all()
    
    return render_template('index.html', 
                         total_students=total_students,
                         today_attendance=today_attendance,
                         recent_records=recent_records)


@main.route('/students')
def students():
    """List all students"""
    all_students = Student.query.order_by(Student.roll_number).all()
    return render_template('students.html', students=all_students)


@main.route('/student/add', methods=['GET', 'POST'])
def add_student():
    """Add a new student"""
    if request.method == 'POST':
        name = request.form.get('name')
        roll_number = request.form.get('roll_number')
        email = request.form.get('email')
        
        if not name or not roll_number:
            flash('Name and Roll Number are required!', 'error')
            return redirect(url_for('main.add_student'))
        
        # Check if roll number already exists
        existing_student = Student.query.filter_by(roll_number=roll_number).first()
        if existing_student:
            flash('Roll number already exists!', 'error')
            return redirect(url_for('main.add_student'))
        
        new_student = Student(name=name, roll_number=roll_number, email=email)
        db.session.add(new_student)
        db.session.commit()
        
        flash(f'Student {name} added successfully!', 'success')
        return redirect(url_for('main.students'))
    
    return render_template('add_student.html')


@main.route('/student/delete/<int:student_id>')
def delete_student(student_id):
    """Delete a student"""
    student = Student.query.get_or_404(student_id)
    db.session.delete(student)
    db.session.commit()
    flash(f'Student {student.name} deleted successfully!', 'success')
    return redirect(url_for('main.students'))


@main.route('/attendance/mark', methods=['GET', 'POST'])
def mark_attendance():
    """Mark attendance for a specific date"""
    if request.method == 'POST':
        attendance_date = request.form.get('date')
        if attendance_date:
            attendance_date = datetime.strptime(attendance_date, '%Y-%m-%d').date()
        else:
            attendance_date = date.today()
        
        students = Student.query.all()
        
        for student in students:
            status = request.form.get(f'status_{student.id}')
            remarks = request.form.get(f'remarks_{student.id}')
            
            if status:
                # Check if attendance already exists for this student and date
                existing = Attendance.query.filter_by(
                    student_id=student.id,
                    date=attendance_date
                ).first()
                
                if existing:
                    existing.status = status
                    existing.remarks = remarks
                else:
                    new_attendance = Attendance(
                        student_id=student.id,
                        date=attendance_date,
                        status=status,
                        remarks=remarks
                    )
                    db.session.add(new_attendance)
        
        db.session.commit()
        flash(f'Attendance marked successfully for {attendance_date}!', 'success')
        return redirect(url_for('main.index'))
    
    # GET request
    selected_date = request.args.get('date')
    if selected_date:
        selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
    else:
        selected_date = date.today()
    
    students = Student.query.order_by(Student.roll_number).all()
    
    # Get existing attendance for the selected date
    existing_attendance = {}
    for student in students:
        record = Attendance.query.filter_by(
            student_id=student.id,
            date=selected_date
        ).first()
        if record:
            existing_attendance[student.id] = record
    
    return render_template('mark_attendance.html', 
                         students=students, 
                         selected_date=selected_date,
                         existing_attendance=existing_attendance)


@main.route('/attendance/view')
def view_attendance():
    """View attendance records with filters"""
    # Get filter parameters
    student_id = request.args.get('student_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Base query
    query = Attendance.query
    
    # Apply filters
    if student_id:
        query = query.filter_by(student_id=student_id)
    if start_date:
        query = query.filter(Attendance.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
    if end_date:
        query = query.filter(Attendance.date <= datetime.strptime(end_date, '%Y-%m-%d').date())
    
    records = query.order_by(Attendance.date.desc()).all()
    students = Student.query.order_by(Student.name).all()
    
    return render_template('view_attendance.html', 
                         records=records, 
                         students=students,
                         selected_student_id=student_id,
                         start_date=start_date,
                         end_date=end_date)


@main.route('/reports')
def reports():
    """Generate attendance reports"""
    report_type = request.args.get('type', 'summary')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Default to current month if no dates provided
    if not start_date or not end_date:
        today = date.today()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    students = Student.query.order_by(Student.roll_number).all()
    
    report_data = []
    for student in students:
        attendance_records = Attendance.query.filter(
            Attendance.student_id == student.id,
            Attendance.date >= start_date_obj,
            Attendance.date <= end_date_obj
        ).all()
        
        total_days = len(attendance_records)
        present_days = sum(1 for r in attendance_records if r.status == 'present')
        absent_days = sum(1 for r in attendance_records if r.status == 'absent')
        late_days = sum(1 for r in attendance_records if r.status == 'late')
        
        percentage = (present_days / total_days * 100) if total_days > 0 else 0
        
        report_data.append({
            'student': student,
            'total_days': total_days,
            'present_days': present_days,
            'absent_days': absent_days,
            'late_days': late_days,
            'percentage': round(percentage, 2)
        })
    
    return render_template('reports.html',
                         report_data=report_data,
                         start_date=start_date,
                         end_date=end_date)


@main.route('/reports/export')
def export_report():
    """Export attendance report to Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        today = date.today()
        start_date = today.replace(day=1).strftime('%Y-%m-%d')
        end_date = today.strftime('%Y-%m-%d')
    
    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Add headers
    headers = ['Roll Number', 'Name', 'Total Days', 'Present', 'Absent', 'Late', 'Attendance %']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    students = Student.query.order_by(Student.roll_number).all()
    
    for row_num, student in enumerate(students, 2):
        attendance_records = Attendance.query.filter(
            Attendance.student_id == student.id,
            Attendance.date >= start_date_obj,
            Attendance.date <= end_date_obj
        ).all()
        
        total_days = len(attendance_records)
        present_days = sum(1 for r in attendance_records if r.status == 'present')
        absent_days = sum(1 for r in attendance_records if r.status == 'absent')
        late_days = sum(1 for r in attendance_records if r.status == 'late')
        percentage = (present_days / total_days * 100) if total_days > 0 else 0
        
        ws.cell(row=row_num, column=1).value = student.roll_number
        ws.cell(row=row_num, column=2).value = student.name
        ws.cell(row=row_num, column=3).value = total_days
        ws.cell(row=row_num, column=4).value = present_days
        ws.cell(row=row_num, column=5).value = absent_days
        ws.cell(row=row_num, column=6).value = late_days
        ws.cell(row=row_num, column=7).value = f"{percentage:.2f}%"
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"attendance_report_{start_date}_to_{end_date}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
